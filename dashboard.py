"""Scheme-driven dashboard renderer. `render(scheme)` draws the whole mobile UI for any
zone-scheme defined in schemes.py."""
import base64
import html
import io
import math
from datetime import datetime, timedelta, timezone
from pathlib import Path

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# Master switch: when True, both dashboards are fully closed — nobody (sales or
# admin) can log in or view anything. Set back to False to reopen.
ACCESS_LOCKED = True

IST = timezone(timedelta(hours=5, minutes=30))  # report dates in India time


def _ordinal(n: int) -> str:
    suf = "th" if 11 <= (n % 100) <= 13 else {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")
    return f"{n:02d}{suf}"


def _till_date() -> str:
    """Yesterday (T-1) in IST, e.g. '03rd June 2026'."""
    y = (datetime.now(IST) - timedelta(days=1)).date()
    return f"{_ordinal(y.day)} {y.strftime('%B %Y')}"

import pandas as pd
import streamlit as st

from schemes import Scheme

KEYS = ["Zone", "State", "Distributor", "Dealer"]
NAVY, ACCENT, GREEN, AMBER, RED = "#003C71", "#F5A623", "#1E8E3E", "#F9A825", "#D93025"
ASSETS = Path(__file__).parent / "assets"
# First match wins: a real PNG/JPG logo takes priority over the SVG placeholder.
LOGO_CANDIDATES = ("jsw_logo.png", "jsw_logo.jpg", "jsw_logo.jpeg", "jsw_logo.svg")
_MIME = {".png": "image/png", ".jpg": "image/jpeg", ".jpeg": "image/jpeg", ".svg": "image/svg+xml"}


def _logo_img() -> str:
    """Return an <img> tag embedding the first logo file found in assets/, or ''."""
    for name in LOGO_CANDIDATES:
        p = ASSETS / name
        if p.exists():
            b64 = base64.b64encode(p.read_bytes()).decode()
            return f'<img class="logo" src="data:{_MIME[p.suffix.lower()]};base64,{b64}" alt="JSW One"/>'
    return ""


# ----------------------------------------------------------------------------
# Data loading (cached; keyed by file paths + newest mtime)
# ----------------------------------------------------------------------------
@st.cache_data
def _load_dated(paths: tuple, mtime: float, eb_iso: str | None) -> pd.DataFrame:
    frames = []
    for p in paths:
        if not Path(p).exists():
            continue
        df = pd.read_csv(p, dtype=str)
        df.columns = [c.strip() for c in df.columns]
        for col in KEYS:
            df[col] = df[col].fillna("").astype(str).str.strip()
        raw = df["Date"].fillna("").astype(str).str.strip()
        iso = raw.str.match(r"^\d{4}[-/.]")
        d = pd.Series(pd.NaT, index=raw.index, dtype="datetime64[ns]")
        d[iso] = pd.to_datetime(raw[iso], errors="coerce", dayfirst=False)
        d[~iso] = pd.to_datetime(raw[~iso], errors="coerce", dayfirst=True)
        df["Date"] = d
        df["MT"] = pd.to_numeric(df["MT"], errors="coerce")
        frames.append(df[(df["Dealer"] != "") & df["MT"].notna()])
    if not frames:
        return pd.DataFrame(columns=KEYS + ["Date", "MT"])
    return pd.concat(frames, ignore_index=True)


@st.cache_data
def _load_agg(paths: tuple, mtime: float) -> pd.DataFrame:
    frames = []
    for p in paths:
        if not Path(p).exists():
            continue
        df = pd.read_csv(p, dtype=str)
        df.columns = [c.strip() for c in df.columns]
        for col in KEYS:
            df[col] = df[col].fillna("").astype(str).str.strip()
        df["MT"] = pd.to_numeric(df["MT"], errors="coerce")
        frames.append(df[(df["Dealer"] != "") & df["MT"].notna()][KEYS + ["MT"]])
    if not frames:
        return pd.DataFrame(columns=KEYS + ["MT"])
    return pd.concat(frames, ignore_index=True)


@st.cache_data
def _load_gifts(path: str, mtime: float) -> pd.DataFrame:
    g = pd.read_csv(path)
    g.columns = [c.strip() for c in g.columns]
    return g.sort_values("MT").reset_index(drop=True)


def _dealer_pieces(scheme: Scheme, mtime: float) -> pd.DataFrame:
    """One row per dealer with WHOLE-MT temporal pieces (half-up rounded), so every
    derived figure is an integer and sums tie out exactly:
      eb   = volume billed on/before the early-bird cutoff (early-bird schemes)
      May  = eb + rest-of-May  ·  June = June volume  ·  Total = eb + May-after + June
    """
    eb_iso = scheme.eb_date.isoformat() if scheme.eb_date else None
    dated = _load_dated(scheme.paths(scheme.dated_files), mtime, eb_iso)
    may_agg = _load_agg(scheme.paths([f for f in scheme.agg_files if "may" in f.lower()]), mtime)
    jun_agg = _load_agg(scheme.paths([f for f in scheme.agg_files if "june" in f.lower()]), mtime)
    cut = pd.Timestamp(scheme.eb_date) if (scheme.early_bird and scheme.eb_date) else None

    def grp(df):
        return df.groupby(KEYS)["MT"].sum() if not df.empty else pd.Series(dtype=float)

    if cut is not None and not dated.empty:
        eb_ex = grp(dated[dated["Date"] <= cut])
        mid_dated = grp(dated[~(dated["Date"] <= cut)])        # May, after the cutoff
    else:
        eb_ex = pd.Series(dtype=float)
        mid_dated = grp(dated) if not dated.empty else pd.Series(dtype=float)
    may_a = grp(may_agg)                                       # May aggregate (non-EB schemes)
    jun_ex = grp(jun_agg)

    idx = None
    for s in (eb_ex, mid_dated, may_a, jun_ex):
        if not s.empty:
            idx = s.index if idx is None else idx.union(s.index)
    if idx is None:
        return pd.DataFrame(columns=KEYS + ["eb", "after", "May", "June", "Total"])

    def ri(s):
        return s.reindex(idx, fill_value=0.0)

    eb_r = ri(eb_ex).map(rhalf)
    mid_r = (ri(mid_dated) + ri(may_a)).map(rhalf)
    jun_r = ri(jun_ex).map(rhalf)

    p = pd.DataFrame(index=idx)
    p["eb"] = eb_r
    p["after"] = mid_r + jun_r
    p["May"] = eb_r + mid_r
    p["June"] = jun_r
    p["Total"] = eb_r + mid_r + jun_r
    return p.reset_index()  # KEYS + eb, after, May, June, Total


def _aggregate(scheme: Scheme, mtime: float) -> pd.DataFrame:
    """One row per dealer with whole-MT scheme metrics."""
    p = _dealer_pieces(scheme, mtime)
    rows = []
    for r in p.itertuples(index=False):
        v_eb, v_after, total = r.eb, r.after, r.Total
        qualified = total >= scheme.min_mt
        if not qualified:
            points, eb_applied = 0.0, False
        elif scheme.early_bird:
            eb_applied = v_eb >= scheme.min_mt
            points = (v_eb * scheme.points_per_mt * (scheme.eb_mult if eb_applied else 1.0)
                      + v_after * scheme.points_per_mt)
        else:
            eb_applied = False
            points = total * scheme.points_per_mt
        rows.append({
            "Zone": r.Zone, "State": r.State, "Distributor": r.Distributor, "Dealer": r.Dealer,
            "MT_eb": int(v_eb), "MT_after": int(v_after),
            "May MT": int(r.May), "June MT": int(r.June),
            "Total MT": int(total), "Points": round(points, 0),
            "Gift Value": round(points * scheme.gift_per_point, 0),
            "Qualified": qualified, "Early Bird": eb_applied and qualified,
        })
    out = pd.DataFrame(rows)
    if not out.empty:
        out = out.sort_values("Total MT", ascending=False).reset_index(drop=True)
    return out


# ----------------------------------------------------------------------------
# Helpers
# ----------------------------------------------------------------------------
def inr(n: float) -> str:
    n = int(round(n)); s = str(abs(n))
    if len(s) > 3:
        last3, rest, parts = s[-3:], s[:-3], []
        while len(rest) > 2:
            parts.insert(0, rest[-2:]); rest = rest[:-2]
        if rest:
            parts.insert(0, rest)
        s = ",".join(parts) + "," + last3
    return ("-" if n < 0 else "") + s


def rhalf(x) -> int:
    """Round half-up to a whole number (10.5 -> 11, 10.4 -> 10)."""
    try:
        return int(math.floor(float(x) + 0.5))
    except (TypeError, ValueError):
        return 0


def mt_fmt(x) -> str:
    """Whole-number MT with thousands separators."""
    return f"{rhalf(x):,}"


def status_of(scheme: Scheme, r) -> tuple[str, str]:
    if not r["Qualified"]:
        return (AMBER, "Almost there") if r["Total MT"] >= 8 else (RED, "Not qualified")
    if scheme.early_bird and r["Early Bird"]:
        return GREEN, "Qualified · Early bird"
    return NAVY, "Qualified"


def gift_for(total_mt, gifts):
    elig = gifts[gifts["MT"] <= total_mt]
    return elig.iloc[-1] if not elig.empty else None


def next_gift(total_mt, gifts):
    nxt = gifts[gifts["MT"] > total_mt]
    return (nxt.iloc[0], nxt.iloc[0]["MT"] - total_mt) if not nxt.empty else (None, 0.0)


def _css():
    st.markdown(f"""
    <style>
      .block-container {{ padding-top: 1.1rem; padding-bottom: 3rem; max-width: 760px; }}
      .hero {{ background: linear-gradient(135deg,{NAVY} 0%,#0A5BA0 100%); color:#fff;
               border-radius:16px; padding:16px 18px; margin-bottom:14px; }}
      .hero-row {{ display:flex; align-items:center; gap:12px; }}
      .hero-txt {{ min-width:0; }}
      .hero .logo {{ height:46px; width:auto; flex:0 0 auto; border-radius:8px;
                     box-shadow:0 1px 3px rgba(0,0,0,.18); }}
      .hero h1 {{ font-size:1.3rem; margin:0; font-weight:800; letter-spacing:.2px; }}
      .hero .sub {{ font-size:.82rem; opacity:.92; margin-top:4px; }}
      .hero .pill {{ display:inline-block; background:rgba(255,255,255,.18); border-radius:999px;
                     padding:2px 10px; font-size:.72rem; font-weight:700; margin-top:8px; }}
      .kpi-grid {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(150px,1fr));
                   gap:10px; margin:6px 0 4px; }}
      .kpi {{ background:#fff; border-radius:14px; padding:13px 14px;
              box-shadow:0 1px 5px rgba(16,24,40,.08); border-left:4px solid {NAVY}; }}
      .kpi .v {{ font-size:1.45rem; font-weight:800; color:{NAVY}; line-height:1.05; }}
      .kpi .l {{ font-size:.70rem; color:#5f6b7a; text-transform:uppercase; letter-spacing:.04em;
                 margin-top:5px; font-weight:600; }}
      .kpi.accent {{ border-left-color:{ACCENT}; }} .kpi.accent .v {{ color:#B9770E; }}
      .kpi.green {{ border-left-color:{GREEN}; }} .kpi.green .v {{ color:{GREEN}; }}
      .row {{ display:flex; align-items:center; gap:10px; background:#fff; border-radius:12px;
              padding:10px 12px; margin-bottom:8px; box-shadow:0 1px 4px rgba(16,24,40,.07); }}
      .row .rank {{ font-weight:800; color:{NAVY}; width:26px; text-align:center; font-size:.95rem; }}
      .row .info {{ flex:1; min-width:0; }}
      .row .name {{ font-weight:700; font-size:.92rem; color:#1a2433; white-space:nowrap;
                    overflow:hidden; text-overflow:ellipsis; }}
      .row .sub {{ font-size:.72rem; color:#6b7686; white-space:nowrap; overflow:hidden;
                   text-overflow:ellipsis; }}
      .row .num {{ text-align:right; font-size:.88rem; font-weight:700; color:#1a2433; }}
      .row .num span {{ display:block; font-size:.70rem; font-weight:600; color:#6b7686; }}
      .dot {{ width:10px; height:10px; border-radius:50%; flex:0 0 10px; }}
      .badge {{ display:inline-block; border-radius:999px; padding:5px 12px; font-size:.8rem; font-weight:700; }}
      .b-green {{ background:#E6F4EA; color:{GREEN}; }}
      .b-amber {{ background:#FEF7E0; color:#9A6700; }}
      .b-red {{ background:#FCE8E6; color:{RED}; }}
      .card {{ background:#fff; border-radius:16px; padding:16px 16px 6px;
               box-shadow:0 2px 8px rgba(16,24,40,.10); margin-bottom:12px; }}
      .card h2 {{ margin:0; font-size:1.15rem; color:{NAVY}; font-weight:800; }}
      .card .csub {{ font-size:.78rem; color:#6b7686; margin:2px 0 10px; }}
      .pbar {{ background:#EAEEF3; border-radius:999px; height:10px; margin:8px 0 4px; overflow:hidden; }}
      .pbar > div {{ height:100%; background:linear-gradient(90deg,{NAVY},#0A5BA0); border-radius:999px; }}
      .ptext {{ font-size:.76rem; color:#6b7686; }}
      .gift {{ background:#FFF8E7; border:1px solid #F4D58B; border-radius:12px; padding:10px 12px;
               margin:10px 0 2px; font-size:.86rem; color:#5a4a1a; }}
      .gift .gcost {{ display:block; margin-top:4px; font-size:.74rem; color:#8a6d2f; font-weight:700; }}
    </style>""", unsafe_allow_html=True)


def _admin_password() -> str:
    try:
        return str(st.secrets["admin_password"])
    except Exception:
        return "east-admin-2026"


def _credentials() -> dict:
    """Team logins as {username: password}. From st.secrets['users'] table, or a
    single app_username/app_password, or a safe fallback."""
    try:
        users = dict(st.secrets["users"])
        if users:
            return {str(k): str(v) for k, v in users.items()}
    except Exception:
        pass
    try:
        return {str(st.secrets.get("app_username", "sales")): str(st.secrets["app_password"])}
    except Exception:
        return {"sales": "powerplay2026"}


def _login_gate(scheme):
    """Single front-page login. The password decides the role:
    team password -> Sales UI; admin password -> Admin UI (costing visible)."""
    if st.session_state.get("auth"):
        return
    st.markdown(
        f'<div class="hero"><div class="hero-row">{_logo_img()}'
        f'<div class="hero-txt"><h1>{scheme.title}</h1>'
        f'<div class="sub">Login</div></div></div></div>',
        unsafe_allow_html=True)
    with st.form("login"):
        u = st.text_input("Username")
        p = st.text_input("Password", type="password")
        ok = st.form_submit_button("Log in")
    if ok:
        creds, admin_pw = _credentials(), _admin_password()
        if p and p == admin_pw:                       # admin password -> admin UI
            st.session_state.update(auth=True, admin=True, user=(u.strip() or "admin"))
            st.rerun()
        elif u in creds and p and p == creds[u]:      # team login -> sales UI
            st.session_state.update(auth=True, admin=False, user=u)
            st.rerun()
        else:
            st.error("Invalid username or password.")
    st.caption("Sales team: use your team username & password. "
               "Admins: enter the admin password to open the admin view.")
    st.stop()


def _sidebar_account():
    """Show who's logged in + role, with a log-out button."""
    with st.sidebar:
        role = "Admin" if st.session_state.get("admin") else "Sales"
        st.markdown(f"👤 **{st.session_state.get('user', '')}** · {role}")
        if st.session_state.get("admin"):
            st.success("Admin mode — costing visible")
        if st.button("Log out"):
            for k in ("auth", "user", "admin"):
                st.session_state.pop(k, None)
            st.rerun()


# ----------------------------------------------------------------------------
# Render
# ----------------------------------------------------------------------------
def render(scheme: Scheme):
    st.set_page_config(page_title=f"Power Play (Q1) · {scheme.region}", page_icon="📊", layout="centered")
    _css()
    if ACCESS_LOCKED:
        st.markdown(
            '<div style="text-align:center; color:#6b7686; font-size:1.1rem; '
            'margin-top:40vh;">offline</div>',
            unsafe_allow_html=True)
        st.stop()
    _login_gate(scheme)

    dated_paths = scheme.paths(scheme.dated_files)
    agg_paths = scheme.paths(scheme.agg_files)
    gifts_path = scheme.paths((scheme.gifts_file,))[0]
    all_paths = [p for p in (*dated_paths, *agg_paths, gifts_path) if Path(p).exists()]
    if not all_paths:
        st.error(f"No data files found for {scheme.region}."); st.stop()
    mtime = max(Path(p).stat().st_mtime for p in all_paths)
    eb_iso = scheme.eb_date.isoformat() if scheme.eb_date else None

    dated = _load_dated(dated_paths, mtime, eb_iso)   # raw billings (for billing history)
    agg = _load_agg(agg_paths, mtime)
    gifts = _load_gifts(gifts_path, mtime)
    dealers = _aggregate(scheme, mtime)
    last_updated = datetime.fromtimestamp(mtime).strftime("%d %b %Y")
    _sidebar_account()
    admin = bool(st.session_state.get("admin"))

    PPM, GPP, MIN = scheme.points_per_mt, scheme.gift_per_point, scheme.min_mt
    EB = scheme.early_bird

    # Hero
    rate = f"₹{int(GPP)}/pt · " if admin else ""
    eb_txt = (f"+25% early-bird by {scheme.eb_date.strftime('%d %b %Y')} · " if EB else "")
    st.markdown(
        f'<div class="hero"><div class="hero-row">{_logo_img()}'
        f'<div class="hero-txt"><h1>{scheme.title}</h1>'
        f'<div class="sub">{int(PPM)} pts/MT · {rate}{eb_txt}qualify at {int(MIN)} MT</div></div></div>'
        f'<span class="pill">Dealer qualifying status · updated till {_till_date()}</span></div>',
        unsafe_allow_html=True)

    if dealers.empty:
        st.info("No dealer data yet."); st.stop()

    # Filters: Zone (only if >1 zone) -> State -> Distributor -> Dealer
    def pick(label, frame, col):
        opts = ["All"] + sorted(frame[col].unique().tolist())
        choice = st.selectbox(label, opts, key=f"f_{col}")
        return frame if choice == "All" else frame[frame[col] == choice]

    multi_zone = dealers["Zone"].nunique() > 1
    fcols = st.columns(4 if multi_zone else 3)
    scope = dealers
    i = 0
    if multi_zone:
        with fcols[i]:
            scope = pick("Zone", scope, "Zone"); i += 1
    with fcols[i]:
        scope = pick("State", scope, "State"); i += 1
    with fcols[i]:
        scope = pick("Distributor", scope, "Distributor"); i += 1
    with fcols[i]:
        scope = pick("Dealer", scope, "Dealer")

    if len(scope) == 1:
        _dealer_detail(scheme, scope.iloc[0], admin, gifts, dated, agg)
    else:
        _kpi_band(scheme, scope, admin)

        # Click-to-view list of qualified dealers (respects the current filters)
        q = scope[scope["Qualified"]].sort_values("Total MT", ascending=False)
        with st.expander(f"🟢 View qualified dealers (≥{int(MIN)} MT) — {len(q)}"):
            if q.empty:
                st.info("No qualified dealers in this scope yet.")
            else:
                qcols = (["Zone"] if multi_zone else []) + \
                    ["Dealer", "Distributor", "State", "Total MT", "Points"]
                if admin:
                    qcols.append("Gift Value")
                qd = q.copy()
                qd["Points"] = qd["Points"].apply(inr)
                if admin:
                    qd["Gift Value"] = qd["Gift Value"].apply(lambda x: "₹" + inr(x))
                st.dataframe(qd[qcols], hide_index=True, width="stretch")
                st.download_button("⬇ Download qualified (CSV)", q[qcols].to_csv(index=False),
                                   file_name=f"{scheme.key}_qualified_dealers.csv",
                                   mime="text/csv", key="dl_qual")

        labels = ["🏆 Leaderboard", "🎯 Push list", "📋 All dealers"]
        if admin:
            labels.append("💰 Summary")
        tabs = st.tabs(labels)
        with tabs[0]:
            top = scope.sort_values(["Points", "Total MT"], ascending=False).head(15)
            st.markdown(_rows(scheme, top, admin, ranked=True), unsafe_allow_html=True)
            eb_leg = "🟢 early bird · " if EB else ""
            st.caption(f"Top 15 by points. {eb_leg}🔵 qualified · 🟠 almost (8–12 MT) · 🔴 not qualified")
        with tabs[1]:
            push = scope[(scope["Total MT"] >= 8) & (scope["Total MT"] < MIN)] \
                .sort_values("Total MT", ascending=False)
            if push.empty:
                st.info("No dealers in the 8–12 MT push range for this scope. 🎉")
            else:
                st.markdown(f"**{len(push)} dealers** are 8–12 MT — one push gets them qualified:")
                st.markdown(_rows(scheme, push, admin, ranked=False, note="push"), unsafe_allow_html=True)
        with tabs[2]:
            show = scope.copy()
            show["Status"] = show.apply(lambda r: status_of(scheme, r)[1], axis=1)
            show["Points"] = show["Points"].apply(inr)
            cols = (["Zone"] if multi_zone else []) + ["Dealer", "Distributor", "State", "Total MT", "Points"]
            if admin:
                show["Qualified Gift"] = show["Total MT"].apply(
                    lambda t: (gift_for(t, gifts)["Gift"] if gift_for(t, gifts) is not None else "—"))
                show["Gift Value"] = show["Gift Value"].apply(lambda x: "₹" + inr(x))
                cols += ["Qualified Gift", "Gift Value"]
            cols.append("Status")
            st.dataframe(show[cols], hide_index=True, width="stretch")
            st.caption("Tip: pick a Dealer in the filter above for full details.")
        if admin:
            with tabs[3]:
                _summary(scheme, scope, gifts, multi_zone)


def _distributor_table(scope, multi_zone) -> pd.DataFrame:
    """Zone × State × Distributor report with May/June/Total/Qualified splits,
    a TOTAL row at the head and per-Zone subtotal rows (screenshot-friendly).
    All volumes come from the per-dealer whole-MT pieces, so May + June = Total."""
    keys = ["Zone", "State", "Distributor"]
    qd = scope[scope["Qualified"]]
    md = scope[scope["May MT"] > 0]
    jd = scope[scope["June MT"] > 0]

    base = scope.groupby(keys).agg(TotDealers=("Dealer", "nunique"),
                                   TotVol=("Total MT", "sum"))
    base["MayDealers"] = md.groupby(keys)["Dealer"].nunique()
    base["MayVol"] = scope.groupby(keys)["May MT"].sum()
    base["JunDealers"] = jd.groupby(keys)["Dealer"].nunique()
    base["JunVol"] = scope.groupby(keys)["June MT"].sum()
    base["QualDealers"] = qd.groupby(keys)["Dealer"].nunique()
    base["QualVol"] = qd.groupby(keys)["Total MT"].sum()
    base = base.fillna(0).reset_index().sort_values(
        ["Zone", "State", "TotVol"], ascending=[True, True, False])

    NUM = ["MayDealers", "MayVol", "JunDealers", "JunVol", "TotDealers", "TotVol", "QualDealers", "QualVol"]
    HEAD = ["# May Dealers", "May Vol (MT)", "# June Dealers", "June Vol (MT)",
            "# Total Dealers Transacted", "Total Vol (MT)", "# Qualified dealers", "Qualified Vol (MT)"]

    def row(zone, state, dist, s):
        d = {"Zone": zone, "State": state, "Distributor Name": dist}
        for h, n in zip(HEAD, NUM):
            d[h] = int(s[n])
        return d

    rows = [row("", "", "TOTAL", base[NUM].sum())]
    for zone in sorted(base["Zone"].unique()):
        zb = base[base["Zone"] == zone]
        for _, r in zb.iterrows():
            rows.append(row(r["Zone"], r["State"], r["Distributor"], r))
        if multi_zone:
            rows.append(row(zone, "", f"{zone} total", zb[NUM].sum()))
    return pd.DataFrame(rows)


def _style_report(df):
    """Bold + shade the TOTAL and zone-subtotal rows; show zeros as '-'."""
    def hl(r):
        d = str(r["Distributor Name"])
        if d == "TOTAL":
            return ["background-color:#DCE6F1; font-weight:700"] * len(r)
        if d.endswith(" total"):
            return ["background-color:#E2EFDA; font-weight:700"] * len(r)
        return [""] * len(r)

    def fmt(v):
        try:
            v = float(v)
        except (TypeError, ValueError):
            return v
        return "-" if v == 0 else f"{v:,.0f}"

    numcols = [c for c in df.columns if c not in ("Zone", "State", "Distributor Name")]
    return df.style.apply(hl, axis=1).format({c: fmt for c in numcols})


def _report_xlsx(report: pd.DataFrame) -> bytes:
    """Build a formatted .xlsx of the report (bold header, shaded TOTAL/zone rows)."""
    from openpyxl.styles import Font, PatternFill, Alignment
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        report.to_excel(w, index=False, sheet_name="Summary")
        ws = w.sheets["Summary"]
        bold = Font(bold=True)
        blue = PatternFill("solid", fgColor="DCE6F1")
        green = PatternFill("solid", fgColor="E2EFDA")
        for c in ws[1]:                              # header
            c.font = bold
            c.alignment = Alignment(wrap_text=True, vertical="center")
        names = list(report["Distributor Name"])
        for i, name in enumerate(names, start=2):
            fill = blue if name == "TOTAL" else green if str(name).endswith(" total") else None
            if fill is not None:
                for c in ws[i]:
                    c.font = bold
                    c.fill = fill
        for col in ws.columns:
            width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max(width + 2, 9), 40)
        ws.freeze_panes = "A2"
    return buf.getvalue()


def _summary(scheme, scope, gifts, multi_zone):
    """Admin report: screenshot-ready Zone/State/Distributor table (May vs June),
    with the gift-wise cost summary tucked into an expander."""
    report = _distributor_table(scope, multi_zone)
    st.markdown("**📦 Zone · State · Distributor — May vs June**")
    st.dataframe(_style_report(report), hide_index=True, width="stretch")
    c1, c2 = st.columns(2)
    c1.download_button("⬇ Excel (.xlsx)", _report_xlsx(report),
                       file_name=f"{scheme.key}_distributor_report.xlsx",
                       mime=XLSX_MIME, key="dl_xlsx", width="stretch")
    c2.download_button("⬇ CSV", report.to_csv(index=False),
                       file_name=f"{scheme.key}_distributor_report.csv",
                       mime="text/csv", key="dl_csv", width="stretch")

    # gift-wise cost summary (kept, but out of the way of the screenshot table)
    df = scope.copy()
    tiers = gifts.sort_values("MT")
    cost_of = {int(r.MT): int(r.Cost) for _, r in tiers.iterrows()}
    name_of = {0: "No gift", **{int(r.MT): f"{int(r.MT)} MT · {r.Gift}" for _, r in tiers.iterrows()}}
    df["TierMT"] = [(int(gift_for(t, gifts)["MT"]) if q and gift_for(t, gifts) is not None else 0)
                    for t, q in zip(df["Total MT"], df["Qualified"])]
    df["GiftCost"] = df["TierMT"].map(lambda m: cost_of.get(m, 0))
    with st.expander(f"🎁 Gift-wise summary · est. running cost ₹{inr(int(df['GiftCost'].sum()))}"):
        grp = df.groupby("TierMT")
        out = pd.DataFrame({
            "Dealers": grp["Dealer"].count(),
            "Total MT": grp["Total MT"].sum(),
            "Est. Cost ₹": grp["GiftCost"].sum(),
        }).reset_index().sort_values("TierMT")
        out["Gift"] = out["TierMT"].map(name_of)
        out["Total MT"] = out["Total MT"].map(mt_fmt)
        out["Est. Cost ₹"] = out["Est. Cost ₹"].map(lambda x: "₹" + inr(x))
        st.dataframe(out[["Gift", "Dealers", "Total MT", "Est. Cost ₹"]],
                     hide_index=True, width="stretch")
        st.caption("Est. cost = dealers × gift catalogue cost (max value, incl. TDS). "
                   "'No gift' = below 12 MT.")


def _kpi_band(scheme, df, admin):
    MIN = scheme.min_mt
    push = int(((df["Total MT"] >= 8) & (df["Total MT"] < MIN)).sum())
    cells = [("", "Active dealers", str(len(df))),
             ("", "Total MT", mt_fmt(df["Total MT"].sum()))]
    if scheme.early_bird:
        cells.append(("", f"MT by {scheme.eb_date.strftime('%d-%b')}", mt_fmt(df["MT_eb"].sum())))
    cells.append(("green", "Qualified ≥12 MT", str(int(df["Qualified"].sum()))))
    if scheme.early_bird:
        cells.append(("green", "Early bird", str(int(df["Early Bird"].sum()))))
    cells.append(("accent", "Push list (8–12)", str(push)))
    cells.append(("", "Total points", inr(df["Points"].sum())))
    if admin:
        cells.append(("accent", "Gift value", "₹" + inr(df["Gift Value"].sum())))
    html_cells = "".join(
        f'<div class="kpi {c}"><div class="v">{v}</div><div class="l">{l}</div></div>'
        for c, l, v in cells)
    st.markdown(f'<div class="kpi-grid">{html_cells}</div>', unsafe_allow_html=True)


def _rows(scheme, df, admin, ranked=True, note=None) -> str:
    out = []
    for i, (_, r) in enumerate(df.iterrows(), 1):
        color, _ = status_of(scheme, r)
        rank = f'<div class="rank">{i}</div>' if ranked else ""
        if note == "push":
            right = f'{mt_fmt(r["Total MT"])} MT<span>needs {mt_fmt(scheme.min_mt - r["Total MT"])} MT</span>'
        else:
            gift = f' · ₹{inr(r["Gift Value"])}' if admin else ""
            right = f'{mt_fmt(r["Total MT"])} MT<span>{inr(r["Points"])} pts{gift}</span>'
        sub = (f'{r["Zone"]} · ' if df["Zone"].nunique() > 1 else "") + \
              f'{html.escape(r["Distributor"])} · {html.escape(r["State"])}'
        out.append(
            f'<div class="row">{rank}<div class="info"><div class="name">{html.escape(r["Dealer"])}</div>'
            f'<div class="sub">{sub}</div></div><div class="num">{right}</div>'
            f'<div class="dot" style="background:{color}"></div></div>')
    return "".join(out)


def _dealer_detail(scheme, r, admin, gifts, dated, agg):
    MIN, PPM, GPP, EB = scheme.min_mt, scheme.points_per_mt, scheme.gift_per_point, scheme.early_bird
    color, label = status_of(scheme, r)
    bcls = {GREEN: "b-green", AMBER: "b-amber", RED: "b-red", NAVY: "b-amber"}[color]
    st.markdown(
        f'<div class="card"><h2>{html.escape(r["Dealer"])}</h2>'
        f'<div class="csub">{html.escape(r["Distributor"])} · {html.escape(r["State"])} · {r["Zone"]}</div>'
        f'<span class="badge {bcls}">{label}</span></div>', unsafe_allow_html=True)
    cols = st.columns(3 if admin else 2)
    cols[0].metric("Total MT", mt_fmt(r["Total MT"]))
    cols[1].metric("Points", inr(r["Points"]))
    if admin:
        cols[2].metric("Gift Value", "₹" + inr(r["Gift Value"]))

    if not r["Qualified"]:
        st.warning(f"Needs **{mt_fmt(MIN - r['Total MT'])} more MT** to reach the {int(MIN)} MT minimum.")
    else:
        ng, need = next_gift(r["Total MT"], gifts)
        if ng is not None:
            tier = ng["MT"]; pct = min(r["Total MT"] / tier, 1.0) * 100
            mult = scheme.eb_mult if (EB and r["Early Bird"]) else 1.0
            tier_pts = tier * PPM * mult
            gnote = f' (₹{inr(tier_pts * GPP)})' if admin else ""
            st.markdown(
                f'<div class="pbar"><div style="width:{pct:.0f}%"></div></div>'
                f'<div class="ptext"><b>{mt_fmt(need)} MT</b> more to the <b>{int(tier)} MT</b> '
                f'milestone → ~{inr(tier_pts)} pts{gnote}</div>', unsafe_allow_html=True)
        else:
            st.success("🏆 Top milestone reached!")

    if EB:
        if r["Early Bird"]:
            st.caption(f"⚡ Early-bird secured: {mt_fmt(r['MT_eb'])} MT by {scheme.eb_date.strftime('%d %b')}.")
        elif r["Qualified"]:
            st.caption(f"No early-bird bonus (only {mt_fmt(r['MT_eb'])} MT by "
                       f"{scheme.eb_date.strftime('%d %b')}; needs ≥{int(MIN)} MT).")

    # qualified gift — admin only
    if admin and r["Qualified"]:
        g = gift_for(r["Total MT"], gifts)
        if g is not None:
            alt = (f' <i>or</i> {html.escape(str(g["AltGift"]))}'
                   if str(g.get("AltGift", "")).strip() else "")
            st.markdown(
                f'<div class="gift">🎁 <b>Qualified gift</b> · {int(g["MT"])} MT tier<br>'
                f'{html.escape(str(g["Gift"]))}{alt}'
                f'<span class="gcost">gift cost ₹{inr(g["Cost"])}</span></div>', unsafe_allow_html=True)
        ng, need = next_gift(r["Total MT"], gifts)
        if ng is not None:
            st.caption(f"➕ {mt_fmt(need)} MT more unlocks the {int(ng['MT'])} MT gift: {ng['Gift']}.")

    # billing history
    sel = lambda d: ((d["Distributor"] == r["Distributor"]) & (d["Dealer"] == r["Dealer"])
                     & (d["State"] == r["State"]) & (d["Zone"] == r["Zone"]))
    hist = []
    if not dated.empty:
        for _, x in dated[sel(dated)].sort_values("Date").iterrows():
            win = ("≤ cutoff (early bird)" if EB and pd.notna(x["Date"]) and x["Date"] <= pd.Timestamp(scheme.eb_date)
                   else "after cutoff" if EB else "")
            hist.append({"Date": x["Date"].strftime("%d %b %Y") if pd.notna(x["Date"]) else "—",
                         "MT": rhalf(x["MT"]), **({"Window": win} if EB else {})})
    amt = agg[sel(agg)]["MT"].sum() if not agg.empty else 0.0
    if amt > 0:
        hist.append({"Date": "Secondary sales (to date)", "MT": rhalf(amt),
                     **({"Window": "after cutoff"} if EB else {})})
    if hist:
        with st.expander(f"Billing detail ({len(hist)} entries)"):
            st.dataframe(pd.DataFrame(hist), hide_index=True, width="stretch")
